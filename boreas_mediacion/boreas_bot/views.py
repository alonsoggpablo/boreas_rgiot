
import os
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json as pyjson
import mimetypes
import tempfile
import openpyxl
from .models import DevicesROUTERS, DevicesNANOENVI, DevicesCO2

# Custom view to display the first 100 rows of each devices* model from the external DB
def devices_external_list(request):
    name_filter = request.GET.get('name', '').strip()
    routers = DevicesROUTERS.objects.all()
    nanoenvi = DevicesNANOENVI.objects.all()
    co2 = DevicesCO2.objects.all()
    if name_filter:
        routers = routers.filter(name__icontains=name_filter)
        nanoenvi = nanoenvi.filter(name__icontains=name_filter)
        co2 = co2.filter(name__icontains=name_filter)
    router_fields = [f.name for f in DevicesROUTERS._meta.fields]
    nanoenvi_fields = [f.name for f in DevicesNANOENVI._meta.fields]
    co2_fields = [f.name for f in DevicesCO2._meta.fields]
    return render(request, 'boreas_bot/devices_external_list.html', {
        'routers': routers,
        'nanoenvi': nanoenvi,
        'co2': co2,
        'router_fields': router_fields,
        'nanoenvi_fields': nanoenvi_fields,
        'co2_fields': co2_fields,
        'name_filter': name_filter,
    })
# boreas_bot/views.py
from django.shortcuts import render
from django.http import HttpResponse
from django.db import connections

def list_devices_tables(request):
    db_conn = connections['default']
    with db_conn.cursor() as cursor:
        cursor.execute("SELECT tablename FROM pg_catalog.pg_tables WHERE tablename ILIKE 'devices%'")
        tables = [row[0] for row in cursor.fetchall()]
    return render(request, 'boreas_bot/devices_tables.html', {'tables': tables})

# Handle JSON file upload and save to json_to_load
@csrf_exempt
def upload_mockup_file(request):
    if request.method == 'POST' and request.FILES.get('file'):
        upload = request.FILES['file']
        # Save to static/json_to_load
        static_dir = os.path.join(settings.BASE_DIR, 'boreas_mediacion', 'static', 'json_to_load')
        os.makedirs(static_dir, exist_ok=True)
        ext = os.path.splitext(upload.name)[1].lower()
        # If JSON, save as-is
        if ext == '.json':
            save_path = os.path.join(static_dir, upload.name)
            with open(save_path, 'wb+') as dest:
                for chunk in upload.chunks():
                    dest.write(chunk)
            # Load JSON data to return
            with open(save_path, 'r', encoding='utf-8') as f:
                try:
                    data = pyjson.load(f)
                except Exception as e:
                    return JsonResponse({'status': 'error', 'message': f'Invalid JSON: {str(e)}'}, status=400)
            return JsonResponse({'status': 'success', 'filename': os.path.basename(save_path), 'data': data})
        # If XLSX, convert to JSON
        elif ext in ['.xlsx', '.xls']:
            # Save to temp file first
            with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
                for chunk in upload.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            try:
                wb = openpyxl.load_workbook(tmp_path)
                ws = wb.active
                data = []
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data.append(dict(zip(headers, row)))
                json_filename = os.path.splitext(upload.name)[0] + '.json'
                save_path = os.path.join(static_dir, json_filename)
                with open(save_path, 'w', encoding='utf-8') as f:
                    pyjson.dump(data, f, ensure_ascii=False, indent=2)
                return JsonResponse({'status': 'success', 'filename': json_filename, 'data': data})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'XLSX to JSON failed: {str(e)}'}, status=400)
            finally:
                os.unlink(tmp_path)
        else:
            return JsonResponse({'status': 'error', 'message': 'Unsupported file type'}, status=400)
    return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)
