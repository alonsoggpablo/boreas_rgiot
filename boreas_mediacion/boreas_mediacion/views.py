import json
from datetime import datetime, date
import paho.mqtt.client as mqtt
import django_filters
from django_filters.rest_framework import DjangoFilterBackend
from django.http import JsonResponse
from paho.mqtt.client import ssl
from rest_framework import viewsets, permissions, generics, status
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes, action
from django.utils import timezone
from django.shortcuts import render
from django.db.models import Max, F, Subquery
from django.views.generic import ListView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import MQTT_topic
from rest_framework.decorators import api_view

from rest_framework.permissions import AllowAny

@api_view(['GET'])
@permission_classes([AllowAny])
def active_mqtt_topics(request):
    """Return all active MQTT topics and their QoS as JSON for the Go client."""
    topics = MQTT_topic.objects.filter(active=True)
    data = [{
        'topic': t.topic,
        'qos': t.qos
    } for t in topics]
    return Response(data)
from .models import reported_measure, MQTT_broker, MQTT_tx, WirelessLogic_SIM, WirelessLogic_Usage, SigfoxDevice, SigfoxReading, MQTT_device_family, DetectedAnomaly
# from .mqtt import client as mqtt_client
from .serializers import (reported_measureSerializer, MQTT_tx_serializer,
                          WirelessLogic_SIMSerializer, WirelessLogic_SIMListSerializer, 
                          WirelessLogic_UsageSerializer, SigfoxDeviceSerializer, SigfoxReadingSerializer,
                          DetectedAnomalySerializer)
from django_filters.rest_framework import DjangoFilterBackend
from .wirelesslogic_service import WirelessLogicService
from django.conf import settings
from rest_framework.authentication import BasicAuthentication
from rest_framework.permissions import IsAuthenticated


# Dashboard Index View
class DashboardIndexView(LoginRequiredMixin, TemplateView):
    template_name = 'dashboard/index.html'
    login_url = '/admin/login/'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get the hostname from the request
        http_host = self.request.META.get('HTTP_HOST', 'localhost:8000')
        # Extract hostname without port
        hostname = http_host.rsplit(':', 1)[0] if ':' in http_host else http_host
        context['airflow_host'] = f'{hostname}:8080'
        return context


# Device Last Reads Dashboard
class DeviceLastReadsView(LoginRequiredMixin, TemplateView):
    template_name = 'dashboard/device_last_reads.html'
    login_url = '/admin/login/'
    
    def get_context_data(self, **kwargs):
        from django.core.paginator import Paginator
        from django.db import connection
        context = super().get_context_data(**kwargs)

        # Get filter parameters from request
        family_filter = self.request.GET.get('family')
        client_filter = self.request.GET.get('client')
        feed_filter = self.request.GET.get('feed')
        page_number = self.request.GET.get('page', 1)

        # PERFORMANCE: Default to most common client if none selected
        # If you want to show all devices, comment out the next two lines
        # if not client_filter:
        #     client_filter = 'Ayuntamiento de Madrid'

        hours = 2

        filters = []
        params = [hours]
        if client_filter:
            filters.append("rm.client = %s")
            params.append(client_filter)
        if family_filter:
            filters.append("f.name = %s")
            params.append(family_filter)
        if feed_filter:
            filters.append("rm.feed = %s")
            params.append(feed_filter)

        where_clause = (" AND " + " AND ".join(filters)) if filters else ""

        query = f"""
            WITH latest_per_device AS (
                SELECT 
                    rm.id,
                    rm.device_id,
                    rm.report_time,
                    rm.name,
                    rm.client,
                    rm.device_family_id_id,
                    rm.feed,
                    f.name as family_name,
                    ROW_NUMBER() OVER (PARTITION BY rm.device_id ORDER BY rm.report_time DESC) as rn
                FROM boreas_mediacion_reported_measure rm
                LEFT JOIN boreas_mediacion_mqtt_device_family f ON rm.device_family_id_id = f.id
                WHERE rm.report_time >= NOW() - INTERVAL '%s hours'{where_clause}
            )
            SELECT id, device_id, report_time, name, client, device_family_id_id, feed, family_name
            FROM latest_per_device
            WHERE rn = 1
            ORDER BY report_time DESC
            LIMIT 200
        """

        with connection.cursor() as cursor:
            cursor.execute(query, params)
            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]

        total_devices = len(results)

        # Get unique families, clients, and feeds for filter dropdowns
        all_families = MQTT_device_family.objects.all().values_list('name', flat=True).order_by('name')
        if family_filter:
            all_feeds = reported_measure.objects.filter(device_family_id__name=family_filter).values_list('feed', flat=True).distinct().order_by('feed')
        else:
            all_feeds = reported_measure.objects.values_list('feed', flat=True).distinct().order_by('feed')
        all_clients = reported_measure.objects.values_list('client', flat=True).distinct().order_by('client')

        context['device_reads'] = results
        context['total_devices'] = total_devices
        context['families'] = all_families
        context['feeds'] = all_feeds
        context['clients'] = all_clients
        context['selected_family'] = family_filter
        context['selected_feed'] = feed_filter
        context['selected_client'] = client_filter
        context['time_window'] = f"Last {hours} hours"

        return context


class reported_measureViewList(generics.ListAPIView):
    queryset = reported_measure.objects.all()
    serializer_class = reported_measureSerializer
    permission_classes = [permissions.AllowAny]  # Permitir acceso público para testing
    filter_backends=[DjangoFilterBackend]
    filterset_fields=['device_id']

class PublishView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    allowed_methods = ['POST', 'GET']
    queryset=MQTT_tx.objects.all()
    serializer_class= MQTT_tx_serializer

    def post(self, request):
        mqtt_server = MQTT_broker.objects.filter(name='rgiot').values_list('server', flat=True)[0]
        mqtt_port = MQTT_broker.objects.filter(name='rgiot').values_list('port', flat=True)[0]
        mqtt_keepalive = MQTT_broker.objects.filter(name='rgiot').values_list('keepalive', flat=True)[0]
        mqtt_user = MQTT_broker.objects.filter(name='rgiot').values_list('user', flat=True)[0]
        mqtt_password = MQTT_broker.objects.filter(name='rgiot').values_list('password', flat=True)[0]
        
        MQTT_SERVER = mqtt_server
        MQTT_PORT = mqtt_port
        MQTT_KEEPALIVE = mqtt_keepalive
        MQTT_USER = mqtt_user
        MQTT_PASSWORD = mqtt_password

        data = request.data

        client = mqtt.Client()
        client.username_pw_set(MQTT_USER, MQTT_PASSWORD)
        client.tls_set(certfile=None,
                       keyfile=None, cert_reqs=ssl.CERT_NONE,
                       tls_version=ssl.PROTOCOL_TLSv1_2, ciphers=None)

        def on_connect(client, userdata, flags, rc):

            if rc == 0:
                print('Connected successfully')
                client.publish(self.request.data['topic'], self.request.data['payload'])
            else:
                print('Bad connection. Code:', rc)
            client.disconnect()

        client.on_connect=on_connect

        client.connect(
            host=MQTT_SERVER,
            port=MQTT_PORT,
            keepalive=MQTT_KEEPALIVE
        )


        client.loop_start()
        return Response({'received data': request.data})

    def get(self, request):
        # Logic for handling GET request
        return Response({"message": "GET request handled"})


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def mqtt_control(request):
    """Control MQTT client: start, stop, or get status"""
    action = request.data.get('action', '')
    # MQTT control is not available; .mqtt module has been removed.
    result = {"status": "error", "message": "MQTT control is not available in this deployment."}
    return Response(result)


# WirelessLogic ViewSets
class WirelessLogic_SIMViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar SIMs de WirelessLogic
    
    Endpoints:
    - list: GET /api/wirelesslogic/sims/
    - retrieve: GET /api/wirelesslogic/sims/{id}/
    - sync_all: POST /api/wirelesslogic/sims/sync_all/
    - sync_usage: POST /api/wirelesslogic/sims/sync_usage/
    """
    queryset = WirelessLogic_SIM.objects.all()
    permission_classes = [permissions.AllowAny]  # Cambiar a IsAuthenticated en producción
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['iccid', 'msisdn', 'status', 'network']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return WirelessLogic_SIMListSerializer
        return WirelessLogic_SIMSerializer
    
    @action(detail=False, methods=['post'])
    def sync_all(self, request):
        """
        Sincroniza todas las SIMs desde la API de WirelessLogic
        
        POST /api/wirelesslogic/sims/sync_all/
        """
        try:
            service = WirelessLogicService()
            created, updated = service.sync_all_sims()
            
            return Response({
                'status': 'success',
                'message': f'Sincronización completada',
                'sims_created': created,
                'sims_updated': updated,
                'total_sims': WirelessLogic_SIM.objects.count()
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def sync_usage(self, request):
        """
        Sincroniza datos de uso de SIMs
        
        POST /api/wirelesslogic/sims/sync_usage/
        Body (opcional): {"days_back": 30}
        """
        try:
            days_back = request.data.get('days_back', 30)
            service = WirelessLogicService()
            usage_count = service.sync_sim_usage(days_back=days_back)
            
            return Response({
                'status': 'success',
                'message': f'Sincronización de uso completada',
                'usage_records_created': usage_count,
                'days_back': days_back
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def usage_history(self, request, pk=None):
        """
        Obtiene historial de uso de una SIM específica
        
        GET /api/wirelesslogic/sims/{id}/usage_history/
        """
        sim = self.get_object()
        usage_records = sim.usage_records.all()
        serializer = WirelessLogic_UsageSerializer(usage_records, many=True)
        return Response(serializer.data)


class WirelessLogic_UsageViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet de solo lectura para datos de uso de SIMs
    
    Endpoints:
    - list: GET /api/wirelesslogic/usage/
    - retrieve: GET /api/wirelesslogic/usage/{id}/
    """
    queryset = WirelessLogic_Usage.objects.all()
    serializer_class = WirelessLogic_UsageSerializer
    permission_classes = [permissions.AllowAny]  # Cambiar a IsAuthenticated en producción
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['sim__iccid', 'period_start', 'period_end']


# Sigfox ViewSets
class SigfoxDeviceViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SigfoxDevice.objects.all()
    serializer_class = SigfoxDeviceSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['device_id']


class SigfoxReadingViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SigfoxReading.objects.all()
    serializer_class = SigfoxReadingSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['device__device_id', 'timestamp']


class SigfoxCallbackView(APIView):
    """Recibe callbacks Sigfox (equivalente al flujo Node-RED /sigfox/gas)."""
    # Ignoramos autenticaciones globales; manejamos Basic auth personalizada dentro del view
    authentication_classes = []
    permission_classes = [permissions.AllowAny]

    def _check_basic_auth(self, request):
        expected = getattr(settings, 'SIGFOX_BASIC_AUTH', None)
        if not expected:
            # valor por defecto rgiot:rgiot codificado
            expected = 'Basic cmdpb3Q6cmdpb3Q='
        received = request.META.get('HTTP_AUTHORIZATION', '')
        return received.strip() == expected

    def post(self, request):
        if not self._check_basic_auth(request):
            return Response({'status': 'error', 'message': 'Unauthorized'}, status=status.HTTP_401_UNAUTHORIZED)

        payload = request.data if isinstance(request.data, dict) else {}
        device_id = payload.get('device')
        data_hex = payload.get('data', '') or ''
        ts = payload.get('timestamp')

        if not device_id or not data_hex:
            return Response({'status': 'error', 'message': 'device y data son obligatorios'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            timestamp = datetime.fromtimestamp(int(ts), tz=timezone.utc) if ts is not None else timezone.now()
        except Exception:
            timestamp = timezone.now()

        # Parseo similar al flow Node-RED
        fw = data_hex[0:1] or None
        temp = None
        hum = None
        co2 = None
        base = None
        try:
            temp = round((int(data_hex[2:6], 16) / 10) - 40, 2)
            hum = int(data_hex[6:8], 16)
            co2 = int(data_hex[8:12], 16)
            base = int(data_hex[12:14], 16)
        except Exception:
            pass

        device, _ = SigfoxDevice.objects.get_or_create(device_id=device_id)
        device.firmware = fw or device.firmware
        device.last_seen = timestamp
        device.last_payload = payload
        device.last_co2 = co2
        device.last_temp = temp
        device.last_hum = hum
        device.last_base = base
        device.save()

        SigfoxReading.objects.create(
            device=device,
            timestamp=timestamp,
            firmware=fw,
            co2=co2,
            temp=temp,
            hum=hum,
            base=base,
            raw_data=payload
        )


# Web Views for Dashboard

def family_last_messages(request):
    from django.contrib.auth.decorators import login_required
    if not request.user.is_authenticated:
        from django.shortcuts import redirect
        return redirect('login')
    family_data = []
    # Add DatadisSupply last read (updated_at) for each active supply
    from .models import DatadisSupply
    latest_supply = DatadisSupply.objects.filter(active=True).order_by('-updated_at').first()
    if latest_supply:
        family_data.append({
            'source': 'datadis_supply',
            'family_name': 'DATADIS Supply',
            'device_id': latest_supply.cups,
            'report_time': latest_supply.updated_at,
            'measures': {
                'address': latest_supply.address,
                'distributor': latest_supply.distributor,
                'point_type': latest_supply.point_type,
            },
        })
    """
    View to display the last message received for each MQTT device family and API data.
    Supports filtering by family name and device_id.
    Shows ALL families, including those without messages.
    """
    # Get all families except 'aemet' for MQTT section
    families = MQTT_device_family.objects.exclude(name__iexact='aemet')
    for family in families:
        last_msg = reported_measure.objects.filter(device_family_id=family.id).order_by('-report_time').first()
        family_data.append({
            'source': 'mqtt',
            'family': family,
            'last_message': last_msg,
            'device_id': last_msg.device_id if last_msg else 'N/A',
            'report_time': last_msg.report_time if last_msg else None,
            'measures': last_msg.measures if last_msg else 'No data',
        })

    # Add latest AEMET API read to API Reads section
    from .models import AemetData
    last_aemet = AemetData.objects.order_by('-timestamp').first()
    if last_aemet:
        family_data.append({
            'source': 'aemet_api',
            'family_name': 'AEMET',
            'device_id': last_aemet.station.station_id,
            'report_time': last_aemet.timestamp,
            'measures': last_aemet.data,
        })
    
    # Add Sigfox data
    sigfox_devices = SigfoxDevice.objects.all().order_by('-updated_at')
    for device in sigfox_devices:
        last_reading = device.readings.first() if hasattr(device, 'readings') else None
        family_data.append({
            'source': 'sigfox',
            'family_name': 'Sigfox',
            'device_id': device.device_id,
            'report_time': device.updated_at,
            'measures': {
                'co2': device.last_co2,
                'temp': device.last_temp,
                'hum': device.last_hum,
                'base': device.last_base,
            },
        })
    
    # ...existing code...
    
    # Add latest WirelessLogic SIM data (only the most recent)
    latest_sim = WirelessLogic_SIM.objects.order_by('-last_sync').first()
    if latest_sim:
        latest_usage = latest_sim.usage_records.order_by('-period_end').first()
        family_data.append({
            'source': 'wirelesslogic',
            'family_name': 'WirelessLogic',
            'device_id': latest_sim.msisdn or latest_sim.iccid,
            'report_time': latest_sim.last_sync,
            'measures': {
                'status': latest_sim.status,
                'network': latest_sim.network,
                'tariff': latest_sim.tariff_name,
                'data_used_mb': latest_usage.data_used_mb if latest_usage else 'N/A',
                'total_cost': latest_usage.total_cost if latest_usage else 'N/A',
            },
        })
    
    # Apply filter if provided
    family_filter = request.GET.get('family_name', '').strip()

    if family_filter:
        family_data = [f for f in family_data if family_filter.lower() == 
                       (f['family'].name.lower() if 'family' in f and f['family'] else 
                        f.get('family_name', '').lower())]
    
    # Sort alphabetically by family/source name
    family_data.sort(key=lambda x: (x['family'].name.lower() if 'family' in x and x['family'] else 
                                    x.get('family_name', '').lower()))
    
    # Calculate total messages
    total_messages = reported_measure.objects.count()
    
    # Get list of all data sources for filter
    all_sources = list(set([f.get('family').name if 'family' in f and f['family'] else 
                           f.get('family_name', '') for f in family_data if f]))
    
    # Extract hostname for Airflow link - check multiple headers
    # Try HTTP_X_FORWARDED_HOST first (set by proxies), then HTTP_HOST, then fallback
    http_host = (request.META.get('HTTP_X_FORWARDED_HOST') or 
                 request.META.get('HTTP_HOST') or 
                 'localhost')
    # Use rsplit with maxsplit=1 to split only on the last colon (port separator)
    hostname = http_host.rsplit(':', 1)[0] if ':' in http_host else http_host
    
    context = {
        'family_data': family_data,
        'families': sorted(all_sources),
        'family_filter': family_filter,
        'total_messages': total_messages,
        'now': timezone.now(),
        'airflow_host': f'{hostname}:8080',
    }
    
    return render(request, 'family_messages.html', context)


# Detected Anomalies API
class DetectedAnomalyViewSet(viewsets.ReadOnlyModelViewSet):
    """
    API endpoint for viewing detected anomalies.
    Read-only as anomalies are created by the Go agent.
    """
    queryset = DetectedAnomaly.objects.all()
    serializer_class = DetectedAnomalySerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['device_id', 'device_name', 'client', 'metric_name', 'anomaly_type', 'detected_at']
    ordering_fields = ['detected_at', 'severity']
    ordering = ['-detected_at']


# External Device Integration ViewSets
    
    search_fields = ['external_device_type_name']
    ordering_fields = ['external_device_type_name', 'created_at']
    ordering = ['external_device_type_name']


    # ExternalDeviceMappingViewSet removed


import os
from django.views.decorators.csrf import csrf_exempt
from django.db import connections
import tempfile
import openpyxl

    # devices_external_list removed (ExternalDeviceMapping)


def list_devices_tables(request):
    """List all device tables in database"""
    db_conn = connections['default']
    with db_conn.cursor() as cursor:
        cursor.execute("SELECT tablename FROM pg_catalog.pg_tables WHERE tablename ILIKE 'devices%'")
        tables = [row[0] for row in cursor.fetchall()]
    return render(request, 'boreas_mediacion/devices_tables.html', {'tables': tables})


@csrf_exempt
def upload_mockup_file(request):
    """Handle JSON/XLSX file upload and conversion"""
    if request.method == 'POST' and request.FILES.get('file'):
        upload = request.FILES['file']
        # Save to static/json_to_load
        static_dir = os.path.join(settings.BASE_DIR, 'boreas_mediacion', 'static', 'json_to_load')
        os.makedirs(static_dir, exist_ok=True)
        ext = os.path.splitext(upload.name)[1].lower()
        # If JSON, save as-is
        if ext == '.json':
            save_path = os.path.join(static_dir, upload.name)
            with open(save_path, 'wb+') as dest:
                for chunk in upload.chunks():
                    dest.write(chunk)
            # Load JSON data to return
            with open(save_path, 'r', encoding='utf-8') as f:
                try:
                    data = json.load(f)
                except Exception as e:
                    return JsonResponse({'status': 'error', 'message': f'Invalid JSON: {str(e)}'}, status=400)
            return JsonResponse({'status': 'success', 'filename': os.path.basename(save_path), 'data': data})
        # If XLSX, convert to JSON
        elif ext in ['.xlsx', '.xls']:
            # Save to temp file first
            with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
                for chunk in upload.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            try:
                wb = openpyxl.load_workbook(tmp_path)
                ws = wb.active
                data = []
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data.append(dict(zip(headers, row)))
                json_filename = os.path.splitext(upload.name)[0] + '.json'
                save_path = os.path.join(static_dir, json_filename)
                with open(save_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                return JsonResponse({'status': 'success', 'filename': json_filename, 'data': data})
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'XLSX to JSON failed: {str(e)}'}, status=400)
            finally:
                os.unlink(tmp_path)
        else:
            return JsonResponse({'status': 'error', 'message': 'Unsupported file type'}, status=400)
    return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)
